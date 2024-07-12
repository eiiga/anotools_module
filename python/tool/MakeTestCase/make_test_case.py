import openpyxl
import pandas as pd


# ファイル名の定数
EXCEL_FILE_PATH = "./input/TestCaseSample.xlsx"
OUTPUT_TESTCASE_PATH = "./output/testcase.csv"

# EXCEL内の定数
SHEET_NAME_MAIN = "main"

COMMON_START_ROW = 4
CASE_NUM_START_ROW = 3
HEADER_START_COL = 2
VALUE_START_COL = 3
CASE_START_COL = 4

DATA_NO_VALUE = "値なし"
TARGET_DATA_SIGN = "○"


# テストケース作成処理
def make_test_case_main():

    # ヘッダー・データ部のリスト初期化
    heder_list = []
    test_data_list = []

    # ヘッダー作成フラグ（True：作成する、False：作成しない）
    is_make_heder = True

    # テストケースのExcelファイル読み込み
    WorkBook = openpyxl.load_workbook(filename=EXCEL_FILE_PATH)

    # ワークシート設定
    WorkSheet = WorkBook[SHEET_NAME_MAIN]

    # 列カウンタ初期化
    i_col = CASE_START_COL

    # [ケースの番号]行が空白になるまで繰り返し
    while not WorkSheet.cell(row=CASE_NUM_START_ROW, column=i_col).value is None:
        
        # 行カウンタ初期化
        i_row = COMMON_START_ROW

        # テストデータ要素リスト初期化
        test_data_element = []

        # [値]行が空白になるまで繰り返し
        while not WorkSheet.cell(row=i_row, column=VALUE_START_COL).value is None:
            
            # ヘッダー作成フラグONならヘッダー情報を取得（ループ1周目目のみ）
            if is_make_heder:
                if not WorkSheet.cell(row=i_row, column=HEADER_START_COL).value is None:
                    # 「データ項目」列が空白でなければヘッダーリストに設定
                    heder_list.append(WorkSheet.cell(row=i_row, column=HEADER_START_COL).value)

            # テストケース対象判定
            if WorkSheet.cell(row=i_row, column=i_col).value == TARGET_DATA_SIGN:

                if WorkSheet.cell(row=i_row, column=VALUE_START_COL).value == DATA_NO_VALUE:
                    # 「値なし」の場合は空白を設定
                    test_data_element.append('')
                else:
                    # 「値」のデータを設定
                    test_data_element.append(WorkSheet.cell(row=i_row, column=VALUE_START_COL).value)

            # 行カウンタインクリメント
            i_row += 1

        # データリストに格納
        test_data_list.append(test_data_element)

        # ヘッダー作成フラグOFF
        is_make_heder = False

        # 列カウンタインクリメント
        i_col += 1
    
    # csv出力
    result_data = pd.DataFrame(test_data_list)
    result_data.to_csv(OUTPUT_TESTCASE_PATH, index=False, header=heder_list)

    # Excelファイルを終了
    WorkBook.close()


# メイン処理
if __name__ == '__main__':
    make_test_case_main()
