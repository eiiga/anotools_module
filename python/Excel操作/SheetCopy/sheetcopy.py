import openpyxl

# ファイルパスの定数
EXCEL_MST_FILEPATH = './mst/mst_sheet.xlsx'
OUTPUT_DIR = './output/'

# シート名の定数
MST_SHEET_MAIN = '本紙'

# 「本紙」シートに関する定数
MST_SHEET_START_ROW = 3
MST_SHEET_FILENAME_COL = 2
MST_SHEET_SHEETNAME_COL = 3

# シートコピー処理
def copy_sheet():
    '''
    ファイル名と指定したシートのみを残す処理
    
    「本紙」シートの「ファイル名」列の値で「mst_sheer.xlsx」を別名保存し
    「シート名」列の値のシート名のみを残す処理
    
    '''
    
    # マスタのExcelファイルを開く
    mst_wb = openpyxl.load_workbook(EXCEL_MST_FILEPATH)

    # シート名を配列に格納
    mst_shtname_list = [ws.title for ws in mst_wb]
    
    # 「本紙」シートがなければ処理終了
    if MST_SHEET_MAIN not in mst_shtname_list:
        print('「本紙」シートがありません')
        
        # マスタExcelファイル閉じる
        mst_wb.close()
        
        return
    
    # 「本紙」シートをセット
    mst_ws = mst_wb[MST_SHEET_MAIN]
    
    # 行カウンタの初期値をセット
    i_row = MST_SHEET_START_ROW
    
    # ファイル名の値が空になるまで繰り返し
    while mst_ws.cell(i_row, MST_SHEET_FILENAME_COL).value != None:
        
        # ファイル名とシート名をセット
        filename = OUTPUT_DIR + mst_ws.cell(i_row, MST_SHEET_FILENAME_COL).value + '.xlsx'
        sheetname = mst_ws.cell(i_row, MST_SHEET_SHEETNAME_COL).value
        
        # 行カウンタをインクリメント
        i_row += 1
        
        # 設定したシート名がなければスキップ
        if sheetname not in mst_shtname_list:
            print('シート名{}がありません'.format(sheetname))
            continue
        
        # 入力したファイル名でマスタExcelファイルを別名保存
        mst_wb.save(filename)
        
        # 別名保存したExcelファイルを開く
        output_wb = openpyxl.load_workbook(filename)
        
        # シート名分繰り返し
        for output_ws in output_wb.worksheets:
            # 入力したシート名以外の場合
            if output_ws.title != sheetname:
                # シートを削除
                output_wb.remove(output_ws)
        
        # 別名保存したExcelを保存
        output_wb.save(filename)
        
        # 別名保存したExcelを閉じる
        output_wb.close()
    
    # マスタExcelファイルを閉じる
    mst_wb.close()


# メイン処理
if __name__ == '__main__':
    # シートコピー処理呼び出し
    copy_sheet()